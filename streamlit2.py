# -*- coding: utf-8 -*-
"""
Created on Wed Jul  2 15:17:04 2025

@author: Andras.Hegedus
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog, simpledialog
from collections import OrderedDict
import pandas as pd
from tkinter import ttk 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
import tkinter as tk



ctk.set_appearance_mode("system")  # vagy "dark" / "light"
ctk.set_default_color_theme("blue")  # vagy "green", "dark-blue", stb.

# ---------- Data Model ---------- #
CATEGORIES = [
    {"name": "Prepare", "scope": "global", "items": [
        {"task": "Desk Study, reconnaissance on satellite images, followed by ground geological survey, identification of problematic zones; Geological mapping at scale 1:2.000 / 1:5.000; Geological survey report; Preparation of a geological survey plan.",
         "unit": "Lumpsum"}
    ]},
    {"name": "Borehole Services", "scope": "borehole", "items": [
        {"task": "Core drillings (m)", "unit": "Meter"},
        {"task": "Backfilling by grouting", "unit": "Meter"},
        {"task": "Installation of PVC piezometer (including water level observation)", "unit": "Piece"},
        {"task": "Monitoring of groundwater in standpipes (bi-monthly over 1 year period incl. reporting)", "unit": "Piece"},
        {"task": "Undisturbed Sampling (UD) with Shelby Tube, incl. delivery to laboratory", "unit": "Per sample"},
        {"task": "Core Sample (from drilling core), incl. delivery to laboratory", "unit": "Per sample"},
        {"task": "Bulk / Bag Sample, incl. delivery to laboratory", "unit": "Per sample"},
        {"task": "Water sample, incl. delivery to laboratory and sample container", "unit": "Per sample"}
    ]},
    {"name": "In-Situ Tests", "scope": "borehole", "items": [
        {"task": "Standard Penetration Test (including evaluation and test report) and taking disturbed sample", "unit": "Meter"},
        {"task": "CPTu (until refusal) (m)", "unit": "Meter"},
        {"task": "Seismic cone penetration test (until refusal) (m)", "unit": "Meter"},
        {"task": "Temperature and Electrical Conductivity Profile", "unit": "Meter"},
        {"task": "Data Logger Installation including one year monitoring and reporting", "unit": "Piece"},
        {"task": "Hydraulic Testing - Short Pumping Test (including evaluation and test report)", "unit": "Per Test"},
    ]},
    {"name": "Soil Mechanical Tests", "scope": "borehole", "items": [
        {"task": "Moisture content", "unit": "Per Test"},
        {"task": "Unit weight", "unit": "Per Test"},
        {"task": "Specific gravity", "unit": "Per Test"},
        {"task": "Grain size distribution by sieve and hydrometer", "unit": "Per Test"},
        {"task": "Atterberg limits", "unit": "Per Test"},
        {"task": "Compaction test (Standard Proctor)", "unit": "Per Test"},
        {"task": "Compaction test (Modified Proctor)", "unit": "Per Test"},
        {"task": "One-Dimensional Consolidation Test", "unit": "Per Test"},
        {"task": "Consolidated-Drained Triaxial Compression Shear Test", "unit": "Per Test"},
        {"task": "Unconsolidated Undrained Triaxial Compression Shear Test", "unit": "Per Test"},
        {"task": "Consolidated Undrained Triaxial Compression Shear Test ", "unit": "Per Test"},
        {"task": "California Bearing Ratio (CBR) - Test", "unit": "Per Test"},
        {"task": "Direct Shear Test (Shear Box Test)", "unit": "Per Test"},
        {"task": "Organic matter content test - bottom part", "unit": "Per Test"},
    ]},
    {"name": "Rock Mechanical Tests", "scope": "borehole", "items": [
        {"task": "Unconfined compressive strength (UCS) - Test (including E-Modulus)", "unit": "Unit"},
        {"task": "Rock Porosity and Bulk Density", "unit": "Unit"},
    ]},
   {"name": "Chemical Tests", "scope": "borehole", "items": [
        {"task": "Chemical analysis of soil / rock (agressiveness of soil towards steel and concrete)", "unit": "Unit"},
        {"task": "Chemical analysis of groundwater (agressiveness of water towards steel and concrete)", "unit": "Unit"},
    ]},
    {"name": "Mobilization & Transport", "scope": "global", "items": [
        {"task": "Mobilisation and demobilisation of key equipment (drilling rig, CPT rig) including ancillary equipment", "unit": "Unit"},
        {"task": "Transport of drilling equipment between investigation sites (within a river crossing / bridge)", "unit": "Piece"},
        {"task": "Transport of drilling equipment between investigation sites (between two river crossings / bridges)", "unit": "Piece"},
    ]},
    {"name": "Geotechnical Report", "scope": "global", "items": [
        {"task": "Preparation of Interpretative Report and piezometer documentation updates", "unit": "Lumpsum"}
    ]},
]

# ---------- Árbetöltő funkció ---------- #
def load_price_map():
    path = filedialog.askopenfilename(title="Open 'Pricelist' excel file", filetypes=[('Excel file', '*.xlsx')])
    if not path:
        return None, None, None
    try:
        df = pd.read_excel(path, sheet_name=0)  # Első sheet
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load the price list sheet: {e}")
        return None, None, None

    price_columns = [col for col in df.columns if col != 'Task']
    if not price_columns:
        messagebox.showerror("Error", "No price column in the Excel.")
        return None, None, None
    
    selected_column = None
    def on_ok():
        nonlocal selected_column
        selected_column = combo.get()
        top.destroy()

    top = tk.Toplevel()
    top.title("Relevant price-column")
    top.resizable(False, False)
    top.configure(bg="#f2f2f2")
    
    # Középre helyezés
    top.update_idletasks()
    width, height = 300, 140
    x = (top.winfo_screenwidth() // 2) - (width // 2)
    y = (top.winfo_screenheight() // 2) - (height // 2)
    top.geometry(f"{width}x{height}+{x}+{y}")
    
    # Címke
    label = tk.Label(top, text="Choose the relevant price column for your project:", font=("Segoe UI", 10), bg="#f2f2f2")
    label.pack(pady=(20, 5))
    
    # Combobox
    combo = ttk.Combobox(top, values=price_columns, state="readonly", font=("Segoe UI", 10))
    combo.pack(pady=5)
    combo.current(0)
    
    # OK gomb
    style = ttk.Style()
    style.configure("Custom.TButton", font=("Segoe UI", 10), padding=6)
    ok_btn = ttk.Button(top, text="OK", command=on_ok, style="Custom.TButton")
    ok_btn.pack(pady=(10, 15))
    
    top.grab_set()
    top.wait_window()
    
    
    if selected_column not in price_columns:
        messagebox.showerror("Error", "Invalid column was choosen.")
        return None, None, None
    
    price_map = df.set_index('Task')[selected_column].to_dict()
    dynamic_price_map = build_dynamic_price_map(df[['Task', selected_column]].rename(columns={selected_column:'Basic'}))
    
    return price_map, selected_column, dynamic_price_map

    
import re

def parse_depth_range(task_str):
    """
    Megpróbálja kiszedni a mélységtartományt pl. '00.00-10.00 meters' -> (0,10)
    Vagy '0 - 15 m' -> (0,15)
    """
    pattern = r'(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)\s*(m|meters)?'
    match = re.search(pattern, task_str)
    if match:
        start = float(match.group(1))
        end = float(match.group(2))
        return start, end
    return None



def build_dynamic_price_map(df):
    """
    Ebből a DataFrame-ből (Task;Basic) összerakjuk a mélység-sávokat.
    A végén explicit módon a kulcs 'Core drillings (m)' lesz.
    """
    dynamic_price_map = {}
    # először gyűjtsük össze minden sorból a sávokat
    for _, row in df.iterrows():
        task = row['Task']
        price = row['Basic']
        m = re.search(r'(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)', task)
        if not m:
            continue
        start, end = float(m.group(1)), float(m.group(2))
        # minden Core drillings
        if 'Core drillings' in task:
            key = 'Core drillings (m)'
        elif 'CPTu' in task:
            key = 'CPTu (until refusal) (m)'
        elif 'Seismic cone penetration test' in task:
            key = 'Seismic cone penetration test (until refusal) (m)'
        else:
            continue


        dynamic_price_map.setdefault(key, []).append((start, end, price))

    # rendezzük
    for k in dynamic_price_map:
        dynamic_price_map[k].sort(key=lambda x: x[0])
    return dynamic_price_map



def core_drilling_pricing(q):
    # q mélység méterben
    q = float(q)
    if q <= 20:
        return q * 35
    elif q <= 40:
        return 20 * 35 + (q - 20) * 40
    else:
        return 20 * 35 + 20 * 40 + (q - 40) * 45

def cptu_pricing(q):
    q = float(q)
    return 30 * min(q, 15) + 40 * max(0, q - 15)

def seismic_cone_pricing(q):
    q = float(q)
    return 35 * min(q, 15) + 45 * max(0, q - 15)

custom_pricing_functions = {
    "Core drillings (m)": core_drilling_pricing,
    "CPTu (until refusal) (m)": cptu_pricing,
    "Seismic cone penetration test (until refusal) (m)": seismic_cone_pricing,
}


# ---------- Pricing függvény generátor ---------- #
def make_pricing_fn(task_name, price_map, dynamic_price_map=None):
    # Priority: custom_pricing_functions > dynamic_price_map > static price_map
    if task_name in custom_pricing_functions:
        return custom_pricing_functions[task_name]

    if dynamic_price_map and task_name in dynamic_price_map:
        price_brackets = dynamic_price_map[task_name]
        def pricing_fn(q):
            try:
                q = float(q)
            except:
                return 0
            total = 0
            remaining = q
            for start, end, unit_price in price_brackets:
                if remaining <= 0:
                    break
                length = min(remaining, end - start)
                total += length * unit_price
                remaining -= length
            return total
        return pricing_fn

    unit_price = price_map.get(task_name, 0)
    def pricing_fn(q):
        try:
            qty = float(q)
        except:
            qty = 0
        return unit_price * qty
    return pricing_fn


def apply_prices_to_categories(price_map, dynamic_price_map=None):
    for cat in CATEGORIES:
        for item in cat['items']:
            task = item['task']
            price = price_map.get(task, None)
            if price is not None or (dynamic_price_map and task in dynamic_price_map):
                item['unit_price'] = price  # Lehet None, ha dinamikus ár van
                item['pricing_fn'] = make_pricing_fn(task, price_map, dynamic_price_map)
            else:
                item['unit_price'] = None
                item['pricing_fn'] = lambda q: 0

# ---------- Export funkció ---------- #
def export_to_excel(general, bores):
    bore_names = [b['entry'].get() for b in bores]
    # Mennyiség és ár oszlopok páronként egymás mellett
    qty_cols = [f'{name} Qty' for name in bore_names]
    price_cols = [f'{name} Price' for name in bore_names]

    cols = ['Task', 'Unit', 'Total Qty', 'Unit Price', 'Total Price'] + qty_cols + price_cols
    rows = []
    bore_sums = {name: 0 for name in bore_names}  # Furásonként összes ár

    # Global sections
    for cat in [c for c in CATEGORIES if c['scope']=='global']:
        for item in cat['items']:
            qty = general.get(item['task'], 0)
            total = item['pricing_fn'](qty)
            row = OrderedDict([('Task', item['task']),
                               ('Unit', item['unit']),
                               ('Total Qty', qty),
                               ('Unit Price', item['unit_price'] or ''),
                               ('Total Price', total)])
            for name in bore_names:
                row[f'{name} Qty'] = ''
                row[f'{name} Price'] = ''
            rows.append(row)
            # Globális tételeket hozzáadjuk az összesítetthez
            global_total = 0
            for cat in [c for c in CATEGORIES if c['scope']=='global']:
                for item in cat['items']:
                    qty = general.get(item['task'], 0)
                    total = item['pricing_fn'](qty)
                    global_total += total
                    
            
    # Borehole sections
    for cat in [c for c in CATEGORIES if c['scope']=='borehole']:
        # Fő kategória sort jelölünk
        rows.append(OrderedDict([('Task', f"-- {cat['name']} --")]))

        # Furásokra vonatkozó sorok
        bore_section_sums = {name: 0 for name in bore_names}

        for item in cat['items']:
            row = OrderedDict([('Task', item['task']),
                               ('Unit', item['unit']),
                               ('Total Qty', ''),
                               ('Unit Price', item['unit_price'] or ''),
                               ('Total Price', '')])
            tot_q = 0
            tot_p = 0
            for b in bores:
                name = b['entry'].get()
                var, qty_ent = b['tasks'].get(item['task'], (None, None))
                if qty_ent:
                    try:
                        q = float(qty_ent.get())
                    except:
                        q = 0
                    row[f"{name} Qty"] = q
                    prc = item['pricing_fn'](q)
                    row[f"{name} Price"] = prc
                    tot_q += q
                    tot_p += prc
                    bore_section_sums[name] += prc
                else:
                    row[f"{name} Qty"] = ''
                    row[f"{name} Price"] = ''
            row['Total Qty'] = tot_q
            row['Total Price'] = tot_p
            rows.append(row)

        # Összegző sor minden furás alatt
        sum_row = OrderedDict([('Task', f"--- Summarized costs ({cat['name']}) ---"),
                               ('Unit', ''),
                               ('Total Qty', ''),
                               ('Unit Price', ''),
                               ('Total Price', '')])
        for name in bore_names:
            sum_row[f"{name} Qty"] = ''
            sum_row[f"{name} Price"] = bore_section_sums[name]
            bore_sums[name] += bore_section_sums[name]
        rows.append(sum_row)

    # Végösszeg sor, az összes furás és globális tételek összesített költsége
    total_sum = global_total + sum(bore_sums.values())
    #total_sum = sum(bore_sums.values())
    total_row = OrderedDict([('Task', '=== Total costs (EUR) ==='),
                             ('Unit', ''),
                             ('Total Qty', ''),
                             ('Unit Price', ''),
                             ('Total Price', total_sum)])
    for name in bore_names:
        total_row[f"{name} Qty"] = ''
        total_row[f"{name} Price"] = bore_sums[name]
    rows.append(total_row)

    df = pd.DataFrame(rows, columns=cols)
    df = df.applymap(lambda x: '' if isinstance(x, (int, float)) and x == 0 else x)

    # Mentés Excel fájlba és formázás
    save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel file', '*.xlsx')])
    if not save_path:
        return

    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Pricing')
        ws = writer.sheets['Pricing']

        # Formázások
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')  # Pasztellkék fejléchez
        qty_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Pasztellsárga mennyiségekhez
        price_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # Pasztellkék árakhoz

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Fejléc formázása
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Vastag kategória sorok és színezés
        for r_idx, row in enumerate(rows, start=2):
            task_val = row['Task']
            if task_val.startswith('-- ') and task_val.endswith(' --'):
                # Kategória sor: vastag betű, szürke háttér
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            elif task_val.startswith('--- Summarized costs'):
                # Összegző sor: vastag betű, világos zöld háttér
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')
            elif task_val.startswith('=== Total'):
                # Teljes összeg: vastag betű, világos narancs háttér
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

            # Mennyiség és ár oszlopok színezése
            for col_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)
                if 'Qty' in col_name:
                    cell.fill = qty_fill
                    cell.alignment = Alignment(horizontal='center')
                elif 'Price' in col_name:
                    cell.fill = price_fill
                    cell.number_format = '€#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        # Oszlopszélességek állítása (szebb megjelenés)
        col_widths = {
            'Task': 45,
            'Unit': 10,
            'Total Qty': 12,
            'Unit Price': 12,
            'Total Price': 14,
        }
        for col in cols:
            width = col_widths.get(col, 12)
            col_letter = get_column_letter(cols.index(col)+1)
            ws.column_dimensions[col_letter].width = width

        # Egységárak formázása a 'Unit Price' oszlopban
        unit_price_col = cols.index('Unit Price') + 1
        for r_idx in range(2, len(rows)+2):
            cell = ws.cell(row=r_idx, column=unit_price_col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '€#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                

    messagebox.showinfo("Done", f"Save of the Excel file was successful: {save_path}")


# ---------- GUI ---------- #
class BoreholePricingApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Bill of Quantities")
        self.geometry("1100x650")

        # Ártérkép betöltése
        self.price_map = None
        self.price_column = None
        self.load_price_map()

        # Globális mennyiségek
        self.general = {}

        # Borehole-ok listája
        self.bores = []

        # Fő keret
        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(fill='both', expand=True, padx=5, pady=5)

        # Felső gombok
        self.top_buttons_frame = ctk.CTkFrame(self.main_frame)
        self.top_buttons_frame.pack(fill='x', pady=5)
        self.btn_add_bore = ctk.CTkButton(self.top_buttons_frame, text="Add new borehole", command=self.add_borehole)
        self.btn_add_bore.pack(side='left', padx=5)
        self.btn_export = ctk.CTkButton(self.top_buttons_frame, text="Export to an Excel file", command=self.export_excel)
        self.btn_export.pack(side='left', padx=5)

        # Ártérkép label
        self.price_map_label = ctk.CTkLabel(self.top_buttons_frame, text=f"Used price type: {self.price_column}")
        self.price_map_label.pack(side='right', padx=10)

        # Általános mennyiségek keret
        self.general_frame = ctk.CTkFrame(self.main_frame)
        self.general_frame.pack(fill='x', pady=5)
        self.general_entries = {}

        # Kategóriák, globális tételek kiírása
        for cat in [c for c in CATEGORIES if c['scope']=='global']:
            lab = ctk.CTkLabel(self.general_frame, text=cat['name'], font=ctk.CTkFont(size=16, weight="bold"))
            lab.pack(anchor='w', pady=(10,0))
            for item in cat['items']:
                frame = ctk.CTkFrame(self.general_frame)
                frame.pack(fill='x', padx=20, pady=2)
                label = ctk.CTkLabel(frame, text=item['task'], width=600, anchor='w')
                label.pack(side='left')
                entry = ctk.CTkEntry(frame, width=100)
                entry.pack(side='left', padx=5)
                self.general_entries[item['task']] = entry

        # Boreholes frame (scrollable)
        self.bores_frame = ctk.CTkScrollableFrame(self.main_frame)
        self.bores_frame.pack(fill='both', expand=True, pady=10)

        # Azonnal hozzáadunk egy furást alapból
        self.add_borehole()

    def load_price_map(self):
        pm, col, dpm = load_price_map()
        if pm is None:
            messagebox.showerror("Error","The program doesn't able to calculate the prices, exit.")
            self.destroy()
            return
        self.price_map = pm
        self.price_column = col
        self.dynamic_price_map = dpm
        
        # Ha van dinamikus ár, kiiktatjuk az egyedi függvényt (pl. core_drilling_pricing)
        for task in ['Core drillings (m)', 'CPTu (until refusal) (m)', 'Seismic cone penetration test (until refusal) (m)']:
            if task in dpm:
                custom_pricing_functions.pop(task, None)

            
        
        apply_prices_to_categories(self.price_map, self.dynamic_price_map)



    def add_borehole(self):
        bore = {}
        frame = ctk.CTkFrame(self.bores_frame)
        frame.pack(fill='x', pady=5, padx=10, anchor='n')
        bore['frame'] = frame

        # Borehole azonosító
        id_frame = ctk.CTkFrame(frame)
        id_frame.pack(fill='x', pady=3)
        ctk.CTkLabel(id_frame, text="Name of the borehole:", width=90).pack(side='left')
        entry = ctk.CTkEntry(id_frame, width=150)
        entry.pack(side='left', padx=3)
        entry.insert(0, f"Bore_{len(self.bores)+1}")
        bore['entry'] = entry

        # Mennyiségek a borehole tételekre
        bore['tasks'] = {}

        # Borehole scope tételek
        for cat in [c for c in CATEGORIES if c['scope']=='borehole']:
            cat_lab = ctk.CTkLabel(frame, text=cat['name'], font=ctk.CTkFont(size=14, weight='bold'))
            cat_lab.pack(anchor='w', pady=(10,0), padx=5)
            for item in cat['items']:
                line = ctk.CTkFrame(frame)
                line.pack(fill='x', padx=20, pady=1)
                task_label = ctk.CTkLabel(line, text=item['task'], width=600, anchor='w')
                task_label.pack(side='left')
                qty_entry = ctk.CTkEntry(line, width=80)
                qty_entry.pack(side='left', padx=5)
                bore['tasks'][item['task']] = (item, qty_entry)

        self.bores.append(bore)

    def export_excel(self):
        # Összegyűjtjük az általános mennyiségeket
        general_qty = {}
        for task, ent in self.general_entries.items():
            try:
                val = float(ent.get())
            except:
                val = 0
            general_qty[task] = val

        export_to_excel(general_qty, self.bores)
      

if __name__ == "__main__":
    app = BoreholePricingApp()
    app.mainloop()
