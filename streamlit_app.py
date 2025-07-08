
# -*- coding: utf-8 -*-
"""
Created on Mon Jul  7 11:56:59 2025

@author: Andras.Hegedus
"""
# -*- coding: utf-8 -*-
"""
Created on Mon Jul 7 11:56:59 2025

@author: Andras.Hegedus
"""

import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


custom_pricing_functions = {}

# ---------- Data Model ---------- #
CATEGORIES = [
    {"name": "Prepare", "scope": "global", "items": [
        {"task": "Desk Study, reconnaissance on satellite images, followed by ground geological survey, identification of problematic zones; Geological mapping at scale 1:2.000 / 1:5.000; Geological survey report; Preparation of a geological survey plan.",
         "unit": "Lumpsum"}
    ]},
    {"name": "Borehole Services", "scope": "borehole", "items": [
        {"task": "Core drillings (m)", "unit": "Meter"},
        {"task": "Backfilling by grouting", "unit": "Meter"},
        {"task": "Installation of PVC piezometer (including water level observation)", "unit": "Piece"},
        {"task": "Monitoring of groundwater in standpipes (bi-monthly over 1 year period incl. reporting)", "unit": "Piece"},
        {"task": "Undisturbed Sampling (UD) with Shelby Tube, incl. delivery to laboratory", "unit": "Per sample"},
        {"task": "Core Sample (from drilling core), incl. delivery to laboratory", "unit": "Per sample"},
        {"task": "Bulk / Bag Sample, incl. delivery to laboratory", "unit": "Per sample"},
        {"task": "Water sample, incl. delivery to laboratory and sample container", "unit": "Per sample"}
    ]},
    {"name": "In-Situ Tests", "scope": "borehole", "items": [
        {"task": "Standard Penetration Test (including evaluation and test report) and taking disturbed sample", "unit": "Meter"},
        {"task": "CPTu (until refusal) (m)", "unit": "Meter"},
        {"task": "Seismic cone penetration test (until refusal) (m)", "unit": "Meter"},
        {"task": "Temperature and Electrical Conductivity Profile", "unit": "Meter"},
        {"task": "Data Logger Installation including one year monitoring and reporting", "unit": "Piece"},
        {"task": "Hydraulic Testing - Short Pumping Test (including evaluation and test report)", "unit": "Per Test"},
    ]},
    {"name": "Soil Mechanical Tests", "scope": "borehole", "items": [
        {"task": "Moisture content", "unit": "Per Test"},
        {"task": "Unit weight", "unit": "Per Test"},
        {"task": "Specific gravity", "unit": "Per Test"},
        {"task": "Grain size distribution by sieve and hydrometer", "unit": "Per Test"},
        {"task": "Atterberg limits", "unit": "Per Test"},
        {"task": "Compaction test (Standard Proctor)", "unit": "Per Test"},
        {"task": "Compaction test (Modified Proctor)", "unit": "Per Test"},
        {"task": "One-Dimensional Consolidation Test", "unit": "Per Test"},
        {"task": "Consolidated-Drained Triaxial Compression Shear Test", "unit": "Per Test"},
        {"task": "Unconsolidated Undrained Triaxial Compression Shear Test", "unit": "Per Test"},
        {"task": "Consolidated Undrained Triaxial Compression Shear Test ", "unit": "Per Test"},
        {"task": "California Bearing Ratio (CBR) - Test", "unit": "Per Test"},
        {"task": "Direct Shear Test (Shear Box Test)", "unit": "Per Test"},
        {"task": "Organic matter content test - bottom part", "unit": "Per Test"},
    ]},
    {"name": "Rock Mechanical Tests", "scope": "borehole", "items": [
        {"task": "Unconfined compressive strength (UCS) - Test (including E-Modulus)", "unit": "Unit"},
        {"task": "Rock Porosity and Bulk Density", "unit": "Unit"},
    ]},
   {"name": "Chemical Tests", "scope": "borehole", "items": [
        {"task": "Chemical analysis of soil / rock (agressiveness of soil towards steel and concrete)", "unit": "Unit"},
        {"task": "Chemical analysis of groundwater (agressiveness of water towards steel and concrete)", "unit": "Unit"},
    ]},
    {"name": "Mobilization & Transport", "scope": "global", "items": [
        {"task": "Mobilisation and demobilisation of key equipment (drilling rig, CPT rig) including ancillary equipment", "unit": "Unit"},
        {"task": "Transport of drilling equipment between investigation sites (within a river crossing / bridge)", "unit": "Piece"},
        {"task": "Transport of drilling equipment between investigation sites (between two river crossings / bridges)", "unit": "Piece"},
    ]},
    {"name": "Geotechnical Report", "scope": "global", "items": [
        {"task": "Preparation of Interpretative Report and piezometer documentation updates", "unit": "Lumpsum"}
    ]},
]
def make_pricing_fn(task_name, price_map, dynamic_price_map=None):
    if task_name in custom_pricing_functions:
        return custom_pricing_functions[task_name]

    if dynamic_price_map and task_name in dynamic_price_map:
        price_brackets = dynamic_price_map[task_name]
        def pricing_fn(q):
            try:
                q = float(q)
            except:
                return 0
            total = 0
            remaining = q
            for start, end, unit_price in price_brackets:
                if remaining <= 0:
                    break
                length = min(remaining, end - start)
                total += length * unit_price
                remaining -= length
            return total
        return pricing_fn

    unit_price = price_map.get(task_name, 0)
    def pricing_fn(q):
        try:
            qty = float(q)
        except:
            qty = 0
        return unit_price * qty
    return pricing_fn


def apply_prices_to_categories(price_map, dynamic_price_map=None):
    for cat in CATEGORIES:
        for item in cat['items']:
            task = item['task']
            price = price_map.get(task, None)
            if price is not None or (dynamic_price_map and task in dynamic_price_map):
                item['unit_price'] = price
                item['pricing_fn'] = make_pricing_fn(task, price_map, dynamic_price_map)
            else:
                item['unit_price'] = None
                item['pricing_fn'] = lambda q: 0


def build_dynamic_price_map(df):
    dynamic_price_map = {}
    for _, row in df.iterrows():
        task = row['Task']
        price = row['Basic']

        m = re.search(r'(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)', task)
        if not m:
            continue
        start, end = float(m.group(1)), float(m.group(2))

        if 'Core drillings' in task:
            key = 'Core drillings (m)'
        elif 'CPTu' in task:
            key = 'CPTu (until refusal) (m)'
        elif 'Seismic cone penetration test' in task:
            key = 'Seismic cone penetration test (until refusal) (m)'
        else:
            continue

        dynamic_price_map.setdefault(key, []).append((start, end, price))

    for k in dynamic_price_map:
        dynamic_price_map[k].sort(key=lambda x: x[0])

    return dynamic_price_map


def load_price_map_streamlit():
    st.header("Load Geotechnical Price List")

    uploaded_file = st.file_uploader("Upload the Excel price list file", type=["xlsx"])
    if not uploaded_file:
        return None, None, None

    try:
        df = pd.read_excel(uploaded_file, sheet_name=0)
    except Exception as e:
        st.error(f"Error reading file: {e}")
        return None, None, None

    if 'Task' not in df.columns:
        st.error("'Task' column not found in the Excel file.")
        return None, None, None

    price_columns = [col for col in df.columns if col != 'Task']
    if not price_columns:
        st.error("No valid price columns found in the Excel file.")
        return None, None, None

    selected_column = st.selectbox("Select the price column for your project:", price_columns)
    if not selected_column:
        st.warning("No price column selected.")
        return None, None, None

    price_map = df.set_index('Task')[selected_column].to_dict()
    dynamic_price_map = build_dynamic_price_map(df[['Task', selected_column]].rename(columns={selected_column: 'Basic'}))

    st.success("Price list loaded successfully!")
    return price_map, selected_column, dynamic_price_map


def export_to_excel_streamlit(general_qty, borehole_data):
    bore_names = list(borehole_data.keys())
    qty_cols = [f'{name} Qty' for name in bore_names]
    price_cols = [f'{name} Price' for name in bore_names]

    cols = ['Task', 'Unit', 'Total Qty', 'Unit Price', 'Total Price'] + qty_cols + price_cols
    rows = []
    bore_sums = {name: 0 for name in bore_names}

    # (scope == 'global')
    global_total = 0
    for cat in [c for c in CATEGORIES if c['scope'] == 'global']:
        for item in cat['items']:
            qty = general_qty.get(item['task'], 0)
            total = item['pricing_fn'](qty) if item.get('pricing_fn') else 0
            global_total += total
            row = OrderedDict([
                ('Task', item['task']),
                ('Unit', item['unit']),
                ('Total Qty', qty),
                ('Unit Price', item.get('unit_price', '') or ''),
                ('Total Price', total)
            ])
            for name in bore_names:
                row[f'{name} Qty'] = ''
                row[f'{name} Price'] = ''
            rows.append(row)

    # for boreholes (scope == 'borehole')
    for cat in [c for c in CATEGORIES if c['scope'] == 'borehole']:
        # Kategória sor
        rows.append(OrderedDict([('Task', f"-- {cat['name']} --")]))
        bore_section_sums = {name: 0 for name in bore_names}

        for item in cat['items']:
            row = OrderedDict([
                ('Task', item['task']),
                ('Unit', item['unit']),
                ('Total Qty', ''),
                ('Unit Price', item.get('unit_price', '') or ''),
                ('Total Price', '')
            ])
            tot_q = 0
            tot_p = 0
            for name in bore_names:
                qty = 0
                if name in borehole_data and item['task'] in borehole_data[name]:
                    qty = borehole_data[name][item['task']] or 0
                row[f"{name} Qty"] = qty if qty != 0 else ''
                price = item['pricing_fn'](qty) if item.get('pricing_fn') else 0
                row[f"{name} Price"] = price if price != 0 else ''
                tot_q += qty
                tot_p += price
                bore_section_sums[name] += price
            row['Total Qty'] = tot_q if tot_q != 0 else ''
            row['Total Price'] = tot_p if tot_p != 0 else ''
            rows.append(row)

        # summary
        sum_row = OrderedDict([
            ('Task', f"--- Summarized costs ({cat['name']}) ---"),
            ('Unit', ''),
            ('Total Qty', ''),
            ('Unit Price', ''),
            ('Total Price', '')
        ])
        for name in bore_names:
            sum_row[f'{name} Qty'] = ''
            sum_row[f'{name} Price'] = bore_section_sums[name]
            bore_sums[name] += bore_section_sums[name]
        rows.append(sum_row)

    # final summa
    total_sum = global_total + sum(bore_sums.values())
    total_row = OrderedDict([
        ('Task', '=== Total costs (EUR) ==='),
        ('Unit', ''),
        ('Total Qty', ''),
        ('Unit Price', ''),
        ('Total Price', total_sum)
    ])
    for name in bore_names:
        total_row[f'{name} Qty'] = ''
        total_row[f'{name} Price'] = bore_sums[name]
    rows.append(total_row)

    df = pd.DataFrame(rows, columns=cols)
    df = df.applymap(lambda x: '' if isinstance(x, (int, float)) and x == 0 else x)

    # excel to buffer
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Pricing')
        ws = writer.sheets['Pricing']

        # formating
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid') 
        qty_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   
        price_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') 
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # header
        for col_idx in range(1, len(cols)+1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # rows
        for r_idx, row in enumerate(rows, start=2):
            task_val = row['Task']
            if task_val.startswith('-- ') and task_val.endswith(' --'):
                # bold, gray background
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                    cell.border = thin_border
            elif task_val.startswith('--- Summarized costs'):
                # bold, green background
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')
                    cell.border = thin_border
            elif task_val.startswith('=== Total'):
                # bold, orange background
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    cell.border = thin_border
            else:
                # normal rows
                for c_idx in range(1, len(cols)+1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border

            # formating
            for c_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if 'Qty' in col_name:
                    cell.fill = qty_fill
                    cell.alignment = Alignment(horizontal='center')
                elif 'Price' in col_name:
                    cell.fill = price_fill
                    cell.number_format = '€#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

        # column width
        col_widths = {
            'Task': 45,
            'Unit': 10,
            'Total Qty': 12,
            'Unit Price': 12,
            'Total Price': 14,
        }
        for col in cols:
            width = col_widths.get(col, 12)
            col_letter = get_column_letter(cols.index(col)+1)
            ws.column_dimensions[col_letter].width = width

        # Unit Price formating
        unit_price_col = cols.index('Unit Price') + 1
        for r_idx in range(2, len(rows)+2):
            cell = ws.cell(row=r_idx, column=unit_price_col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '€#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    buffer.seek(0)
    return buffer


def main():
    st.title("Bill of Quantities")

    price_map, selected_price_col, dynamic_price_map = load_price_map_streamlit()
    if price_map is None:
        st.info("Please upload a price list to continue.")
        return

    apply_prices_to_categories(price_map, dynamic_price_map)

    st.header("Input Quantities")

    general_qty = {}
    for cat in CATEGORIES:
        if cat["scope"] == "global":
            st.subheader(f"{cat['name']} (Global)")
            for item in cat["items"]:
                qty = st.number_input(f"{item['task']} [{item['unit']}]", min_value=0.0, step=1.0, key=f"general_{item['task']}")
                general_qty[item['task']] = qty

    boreholes = st.text_input("Enter borehole IDs separated by commas", "BH1, BH2")
    borehole_list = [bh.strip() for bh in boreholes.split(",") if bh.strip()]
    borehole_data = {}
    for bh in borehole_list:
        st.subheader(f"Borehole: {bh}")
        borehole_data[bh] = {}
        for cat in CATEGORIES:
            if cat["scope"] == "borehole":
                st.markdown(f"**{cat['name']}**")
                for item in cat["items"]:
                    qty = st.number_input(f"{item['task']} [{item['unit']}] ({bh})", min_value=0.0, step=1.0, key=f"{bh}_{item['task']}")
                    borehole_data[bh][item['task']] = qty

    # Show calculated prices
    st.header("Summary of Costs")
    total_cost = 0

    st.subheader("Global Tasks")
    for task, qty in general_qty.items():
        # Find pricing fn
        pricing_fn = None
        for cat in CATEGORIES:
            for item in cat['items']:
                if item['task'] == task:
                    pricing_fn = item.get('pricing_fn')
                    break
        price = pricing_fn(qty) if pricing_fn else 0
        st.write(f"{task}: Quantity: {qty}, Cost: {price:.2f} EUR")
        total_cost += price

    st.subheader("Borehole Tasks")
    for bh, tasks in borehole_data.items():
        st.markdown(f"**Borehole {bh}**")
        for task, qty in tasks.items():
            pricing_fn = None
            for cat in CATEGORIES:
                for item in cat['items']:
                    if item['task'] == task:
                        pricing_fn = item.get('pricing_fn')
                        break
            price = pricing_fn(qty) if pricing_fn else 0
            st.write(f"{task}: Quantity: {qty}, Cost: {price:.2f} EUR")
            total_cost += price

    st.header(f"Total Cost: {total_cost:.2f} EUR")

    # Export button
    if st.button("Export Bill of Quantities to Excel"):
        excel_buffer = export_to_excel_streamlit(general_qty, borehole_data)
        st.download_button(
            label="Download Excel File",
            data=excel_buffer,
            file_name="Geotechnical_Bill_of_Quantities.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


if __name__ == "__main__":
    main()
